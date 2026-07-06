import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime
import re
import os
import numpy as np
import cv2
import pytesseract
from PIL import Image

st.set_page_config(
    page_title="Tools Rekap - CSV & Ekstraksi Foto",
    page_icon="📄",
    layout="wide"
)

# =========================
# FORMAT BULAN INDONESIA
# =========================
bulan_id = {
    1: "JANUARI", 2: "FEBRUARI", 3: "MARET", 4: "APRIL",
    5: "MEI", 6: "JUNI", 7: "JULI", 8: "AGUSTUS",
    9: "SEPTEMBER", 10: "OKTOBER", 11: "NOVEMBER", 12: "DESEMBER"
}

def format_tanggal_indonesia(tanggal_str):
    dt = datetime.strptime(tanggal_str, "%d-%m-%Y")
    return f"{dt.day} {bulan_id[dt.month]} {dt.year}"


def _is_no_column(col_name: str) -> bool:
    """Deteksi apakah nama kolom merujuk ke kolom 'No' / nomor urut
    (fleksibel: 'No', 'No.', 'NO', 'No Urut', 'Nomor', dll)."""
    normalized = re.sub(r"[^a-z0-9]", "", str(col_name).lower())
    return normalized in ("no", "nourut", "nomor", "nomorurut")


def clean_excess_spaces(col: pd.Series) -> pd.Series:
    """Rapikan spasi berlebihan (mis. 'SJ   0001' -> 'SJ 0001', '  1  ' -> '1')
    pada kolom object, tanpa mengubah nilai kosong/NaN."""
    if col.dtype != "object":
        return col
    return col.apply(
        lambda v: re.sub(r"\s+", " ", str(v)).strip() if pd.notna(v) else v
    )


def try_convert_numeric(col: pd.Series) -> pd.Series:
    """
    Kolom bertipe object (string) dicoba dikonversi jadi numerik (kuantitas),
    supaya di Excel tersimpan sebagai angka (int/float), bukan teks.

    Aturan:
    - Mendukung format angka Indonesia (mis. "1.234,50" -> 1234.50)
      maupun format polos (mis. "1234" -> 1234).
    - Kolom hanya dikonversi jika SELURUH nilai (non-kosong) berhasil
      diparse sebagai angka. Kalau ada satu saja nilai yang bukan angka
      (nama, kode berhuruf, alamat, dll), kolom dibiarkan tetap string.
    - Kalau semua nilai berupa bilangan bulat -> dikonversi ke Int64
      (integer nullable, tetap aman kalau ada sel kosong/NaN).
    - Kalau ada nilai desimal -> dikonversi ke float, tetap numerik
      (bukan string), bukan int.
    """
    if col.dtype != "object":
        return col

    original = col.copy()
    stripped = col.astype(str).str.strip()

    # Anggap sel kosong sebagai NaN (tidak menggagalkan deteksi numerik)
    is_blank = stripped.eq("") | stripped.str.lower().eq("nan")

    # Bersihkan format angka Indonesia: "1.234,50" -> "1234.50"
    cleaned = stripped.str.replace(".", "", regex=False)
    cleaned = cleaned.str.replace(",", ".", regex=False)

    numeric_col = pd.to_numeric(cleaned, errors="coerce")

    # Jika ada nilai non-blank yang gagal diparse -> bukan kolom numerik,
    # kembalikan kolom aslinya (string) tanpa perubahan.
    if numeric_col[~is_blank].isna().any():
        return original

    # Semua nilai non-blank berhasil jadi angka
    if (numeric_col.dropna() % 1 == 0).all():
        return numeric_col.astype("Int64")
    return numeric_col


# =====================================================================
# MENU 1: CSV -> XLSX CONVERTER
# =====================================================================
def render_csv_to_xlsx():
    st.title("📄 Convert CSV ke XLSX")
    st.write("Upload file CSV lalu convert ke Excel (.xlsx)")

    uploaded_file = st.file_uploader(
        "Upload File CSV",
        type=["csv"],
        key="csv_uploader"
    )

    today = datetime.now().strftime("%d-%m-%Y")
    original_filename = ""
    suggested_name = f"template_{today}"

    if uploaded_file is not None:
        original_filename = uploaded_file.name
        suggested_name = os.path.splitext(original_filename)[0]

    st.subheader("Pengaturan Nama File")

    save_mode = st.radio(
        "Pilih metode penyimpanan:",
        ["Custom", "Template"],
        horizontal=True,
        key="save_mode_csv"
    )

    if save_mode == "Custom":
        custom_name = st.text_input(
            "Masukkan nama file",
            value=suggested_name,
            key="custom_name_csv"
        )
        if custom_name.strip():
            file_name = f"{custom_name.strip()}.xlsx"
        else:
            file_name = f"{suggested_name}.xlsx"
    else:
        col1, col2, col3 = st.columns(3)

        with col1:
            jenis_dokumen = st.selectbox(
                "Jenis Dokumen",
                ["PENJUALAN", "INVOICE"],
                key="jenis_dokumen_csv"
            )
        with col2:
            kategori1 = st.selectbox(
                "Produk",
                ["OLI", "LPG"],
                key="kategori1_csv"
            )
        with col3:
            kategori2 = st.selectbox(
                "Lokasi",
                ["SRB", "SGE"],
                key="kategori2_csv"
            )

        now = datetime.now()
        tanggal_text = f"{now.day} {bulan_id[now.month]} {now.year}"

        if uploaded_file is not None:
            dates = re.findall(r"\d{2}-\d{2}-\d{4}", uploaded_file.name)

            if len(dates) >= 2:
                tgl_awal = format_tanggal_indonesia(dates[0])
                tgl_akhir = format_tanggal_indonesia(dates[1])
                if dates[0] == dates[1]:
                    tanggal_text = tgl_awal
                else:
                    tanggal_text = f"{tgl_awal} - {tgl_akhir}"
            elif len(dates) == 1:
                tanggal_text = format_tanggal_indonesia(dates[0])

        file_name = f"{jenis_dokumen} {kategori1} {kategori2} {tanggal_text}.xlsx"

    st.write(f"**Nama file yang akan disimpan:** `{file_name}`")

    if uploaded_file is not None:
        st.info(
            'Jika nama mengandung koma seperti:\n'
            '"PT MAJU, JAYA"\n'
            "pastikan di CSV dibungkus tanda kutip."
        )

        try:
            df = pd.read_csv(
                uploaded_file,
                sep=",",
                quotechar='"',
                encoding="utf-8",
                engine="python"
            )

            # 1) Trim spasi di semua kolom teks
            df = df.apply(
                lambda col: col.str.strip() if col.dtype == "object" else col
            )

            # 2) Rapikan spasi berlebihan (spasi ganda, dll) khusus kolom "No"
            for c in df.columns:
                if _is_no_column(c):
                    df[c] = clean_excess_spaces(df[c])

            # 3) Kolom kuantitas/angka dikonversi ke numerik (int/float),
            #    bukan disimpan sebagai string
            df = df.apply(try_convert_numeric)

            cols = list(df.columns)
            if len(cols) >= 2:
                cols[0], cols[1] = cols[1], cols[0]
                df = df[cols]

            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Data")
            output.seek(0)

            st.success("CSV berhasil dibaca")

            st.download_button(
                label="⬇ Download XLSX",
                data=output,
                file_name=file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_csv_xlsx"
            )

            st.subheader("Preview Data")
            st.dataframe(df, use_container_width=True)

        except Exception as e:
            st.error(f"Gagal membaca file CSV: {e}")


# =====================================================================
# MENU 2: EKSTRAKSI TABEL DARI FOTO (OCR LOKAL - TANPA AI/ANTHROPIC)
# =====================================================================
# Pipeline ini 100% jalan lokal di komputer (OpenCV + Tesseract OCR),
# tidak memanggil API apapun, tidak butuh API key, tidak butuh internet.
# Karena tidak pakai AI vision, akurasinya lebih rendah dibanding model AI -
# terutama untuk kolom angka dan tabel dengan header bertingkat. Hasil
# ekstraksi WAJIB diperiksa & dikoreksi manual sebelum dipakai.

COLUMNS = [
    ("tanggal", "Tanggal"),
    ("no_sj", "No. SJ/DO"),
    ("sopir", "Sopir"),
    ("nama_supplier", "Nama Supplier"),
    ("kg50_is", "50KG - IS"),
    ("kg50_ksg", "50KG - KSG"),
    ("kg12_is", "12KG - IS"),
    ("kg12_ksg", "12KG - KSG"),
    ("kg55_is", "5,5KG - IS"),
    ("kg55_ksg", "5,5KG - KSG"),
    ("istgb_50", "IS+TGB 50KG"),
    ("istgb_12", "IS+TGB 12KG"),
    ("istgb_55", "IS+TGB 5,5KG"),
    ("is_total", "Baris Total?"),
]

# Label kolom data (tanpa "tanggal" dan "is_total"), urutan kiri->kanan
# sesuai struktur tabel LPG standar. Dipetakan sesuai jumlah kolom yang
# berhasil dideteksi dari grid foto.
DATA_LABELS_ORDERED = [label for key, label in COLUMNS if key not in ("tanggal", "is_total")]


def _get_line_positions(sum_arr, min_gap=8, thresh_ratio=0.15):
    """Cari posisi garis (baris/kolom) dari profil proyeksi piksel putih."""
    if sum_arr.max() == 0:
        return []
    thresh = sum_arr.max() * thresh_ratio
    idx = np.where(sum_arr > thresh)[0]
    if len(idx) == 0:
        return []
    groups = []
    start = idx[0]
    prev = idx[0]
    for i in idx[1:]:
        if i - prev > min_gap:
            groups.append((start + prev) // 2)
            start = i
        prev = i
    groups.append((start + prev) // 2)
    return groups


def _clean_lines(lines, min_gap=15):
    """Gabungkan garis yang terlalu berdekatan (noise) jadi satu."""
    if not lines:
        return lines
    cleaned = [lines[0]]
    for v in lines[1:]:
        if v - cleaned[-1] < min_gap:
            continue
        cleaned.append(v)
    return cleaned


def _detect_grid(gray):
    """Deteksi garis horizontal & vertikal tabel dari gambar grayscale."""
    thresh = cv2.adaptiveThreshold(
        ~gray, 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY, 15, -2
    )
    h, w = gray.shape

    horizontal = thresh.copy()
    h_size = max(10, w // 30)
    h_struct = cv2.getStructuringElement(cv2.MORPH_RECT, (h_size, 1))
    horizontal = cv2.erode(horizontal, h_struct)
    horizontal = cv2.dilate(horizontal, h_struct)

    vertical = thresh.copy()
    v_size = max(10, h // 30)
    v_struct = cv2.getStructuringElement(cv2.MORPH_RECT, (1, v_size))
    vertical = cv2.erode(vertical, v_struct)
    vertical = cv2.dilate(vertical, v_struct)

    row_sum = horizontal.sum(axis=1) / 255
    col_sum = vertical.sum(axis=0) / 255

    ys = _clean_lines(_get_line_positions(row_sum), min_gap=15)
    xs = _get_line_positions(col_sum)
    xs_filtered = [xs[0]] if xs else []
    for v in xs[1:]:
        if v - xs_filtered[-1] < 25:
            continue
        xs_filtered.append(v)
    return xs_filtered, ys


def _ocr_cell(gray, y1, y2, x1, x2, numeric=False, pad=4):
    """OCR satu sel tabel. numeric=True membatasi karakter ke angka saja."""
    y1p = max(0, y1 + pad)
    y2p = max(y1p + 1, y2 - pad)
    x1p = max(0, x1 + pad)
    x2p = max(x1p + 1, x2 - pad)
    crop = gray[y1p:y2p, x1p:x2p]
    if crop.size == 0:
        return ""
    crop = cv2.resize(crop, None, fx=4, fy=4, interpolation=cv2.INTER_CUBIC)
    crop = cv2.GaussianBlur(crop, (3, 3), 0)
    _, bw = cv2.threshold(crop, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    bw = cv2.medianBlur(bw, 3)

    if numeric:
        cfg = "--psm 7 -c tessedit_char_whitelist=0123456789-,."
    else:
        cfg = "--psm 6"
    txt = pytesseract.image_to_string(bw, config=cfg)
    txt = re.sub(r"[^A-Za-z0-9,./\-\s]", "", txt).strip()
    txt = re.sub(r"\s+", " ", txt)
    return txt


def _extract_title_date(gray):
    """Coba baca teks judul (baris paling atas gambar) untuk cari tanggal."""
    h, w = gray.shape
    top_strip = gray[0: int(h * 0.14), :]
    txt = pytesseract.image_to_string(top_strip, config="--psm 6")
    match = re.search(
        r"(\d{1,2}\s+(?:Jan|Feb|Mar|Apr|Mei|Jun|Jul|Agu|Sep|Okt|Nov|Des)\w*\s+\d{4})",
        txt, flags=re.IGNORECASE
    )
    return match.group(1) if match else ""


def extract_table_local(image_bytes):
    """
    Ekstrak tabel dari foto secara lokal (OpenCV + Tesseract).
    Return: (tanggal_terdeteksi, list_of_row_dicts (label -> value))
    """
    file_bytes = np.frombuffer(image_bytes, np.uint8)
    img = cv2.imdecode(file_bytes, cv2.IMREAD_COLOR)
    if img is None:
        raise ValueError("Gagal membaca file gambar.")

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    xs, ys = _detect_grid(gray)

    if len(xs) < 3 or len(ys) < 3:
        raise ValueError(
            "Garis tabel tidak terdeteksi dengan jelas pada foto ini. "
            "Coba foto ulang dengan pencahayaan lebih rata dan tabel tidak miring."
        )

    row_heights = [ys[i + 1] - ys[i] for i in range(len(ys) - 1)]
    median_h = sorted(row_heights)[len(row_heights) // 2] if row_heights else 20

    n_cols = len(xs) - 1
    col_labels = DATA_LABELS_ORDERED[:n_cols]
    while len(col_labels) < n_cols:
        col_labels.append(f"Kolom Ekstra {len(col_labels) + 1}")

    tanggal = _extract_title_date(gray)
    rows = []

    for i in range(len(ys) - 1):
        y1, y2 = ys[i], ys[i + 1]
        if y2 - y1 > median_h * 1.8:
            # Baris tinggi = kemungkinan blok header, lewati (bukan data)
            continue
        row = {}
        for j in range(n_cols):
            x1, x2 = xs[j], xs[j + 1]
            is_numeric_col = j >= 3  # 3 kolom pertama = No SJ/Sopir/Supplier
            row[col_labels[j]] = _ocr_cell(gray, y1, y2, x1, x2, numeric=is_numeric_col)
        # deteksi baris TOTAL/GRAND TOTAL dari teks nama supplier
        supplier_text = row.get("Nama Supplier", "").upper()
        row["is_total"] = "YA" if ("TOTAL" in supplier_text) else ""
        rows.append(row)

    return tanggal, rows, col_labels


def build_excel_bytes(rows_df):
    """Bangun file excel (BytesIO) dari dataframe hasil ekstraksi/koreksi."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap LPG"

    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    total_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = list(rows_df.columns)
    if "Baris Total?" in headers:
        headers.remove("Baris Total?")

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for _, row in rows_df.iterrows():
        is_total = str(row.get("Baris Total?", "")).strip().upper() == "YA"
        values = [row.get(h, "") for h in headers]
        ws.append(values)
        r = ws.max_row
        for c in range(1, len(values) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
            if is_total:
                cell.fill = total_fill
                cell.font = Font(bold=True)

    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16
    if "Nama Supplier" in headers:
        ws.column_dimensions[get_column_letter(headers.index("Nama Supplier") + 1)].width = 24
    ws.freeze_panes = "A2"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def render_photo_extraction():
    st.title("📸 Ekstrak Tabel dari Foto (OCR Lokal)")
    st.warning(
        "Fitur ini berjalan 100% lokal di komputer (OpenCV + Tesseract OCR), "
        "**tidak memakai API/AI apapun** — tidak butuh API key dan tidak butuh internet. "
        "Karena tidak pakai AI vision, akurasinya lebih rendah, terutama untuk kolom "
        "angka dan tabel dengan header bertingkat. **Selalu periksa & koreksi hasil "
        "sebelum export**, gunakan foto pratinjau di bawah untuk mencocokkan."
    )

    if "rows_data" not in st.session_state:
        st.session_state.rows_data = []
    if "columns_order" not in st.session_state:
        st.session_state.columns_order = None

    uploaded_photos = st.file_uploader(
        "Upload Foto Tabel (bisa lebih dari satu)",
        type=["jpg", "jpeg", "png", "webp"],
        accept_multiple_files=True,
        key="photo_uploader"
    )

    if uploaded_photos:
        with st.expander("🖼 Pratinjau foto (untuk mencocokkan hasil OCR)", expanded=False):
            preview_cols = st.columns(min(3, len(uploaded_photos)))
            for i, photo in enumerate(uploaded_photos):
                photo.seek(0)
                preview_cols[i % len(preview_cols)].image(photo, caption=photo.name, use_container_width=True)
                photo.seek(0)

    col_a, col_b, col_c = st.columns(3)
    with col_a:
        process_clicked = st.button("🔎 Proses OCR", key="process_btn", use_container_width=True)
    with col_b:
        add_blank_clicked = st.button("➕ Tambah Baris Kosong", key="add_blank_btn", use_container_width=True)
    with col_c:
        clear_clicked = st.button("🧹 Bersihkan Semua Hasil", key="clear_btn", use_container_width=True)

    if clear_clicked:
        st.session_state.rows_data = []
        st.session_state.columns_order = None
        st.rerun()

    if add_blank_clicked:
        labels = st.session_state.columns_order or [l for _, l in COLUMNS]
        blank = {label: "" for label in labels}
        st.session_state.rows_data.append(blank)

    if process_clicked:
        if not uploaded_photos:
            st.warning("Upload minimal satu foto dulu.")
        else:
            progress = st.progress(0, text="Memulai proses OCR...")
            total = len(uploaded_photos)
            for idx, photo in enumerate(uploaded_photos, start=1):
                progress.progress(idx / total, text=f"Memproses foto {idx}/{total}: {photo.name}")
                try:
                    photo.seek(0)
                    image_bytes = photo.read()
                    tanggal, rows, col_labels = extract_table_local(image_bytes)

                    full_labels = ["Tanggal"] + col_labels + ["Baris Total?"]
                    if st.session_state.columns_order is None:
                        st.session_state.columns_order = full_labels

                    for r in rows:
                        row_out = {"Tanggal": tanggal}
                        for label in col_labels:
                            row_out[label] = r.get(label, "")
                        row_out["Baris Total?"] = r.get("is_total", "")
                        st.session_state.rows_data.append(row_out)

                except Exception as e:
                    st.error(f"Gagal memproses {photo.name}: {e}")

            progress.progress(1.0, text="Selesai!")
            st.success(
                f"Selesai memproses {total} foto. Total baris: {len(st.session_state.rows_data)}. "
                "Silakan periksa dan koreksi hasil di tabel bawah."
            )

    st.subheader("Hasil Ekstraksi (WAJIB diperiksa & dikoreksi langsung di tabel)")

    if st.session_state.rows_data:
        display_labels = st.session_state.columns_order or [label for _, label in COLUMNS]
        df = pd.DataFrame(st.session_state.rows_data)
        for label in display_labels:
            if label not in df.columns:
                df[label] = ""
        df = df[display_labels]

        edited_df = st.data_editor(
            df,
            use_container_width=True,
            num_rows="dynamic",
            key="data_editor_rows"
        )

        st.session_state.rows_data = edited_df.to_dict("records")

        excel_bytes = build_excel_bytes(edited_df)
        st.download_button(
            label="💾 Export ke Excel",
            data=excel_bytes,
            file_name="Rekapitulasi_LPG.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_excel_lpg"
        )
    else:
        st.info("Belum ada data. Upload foto lalu klik 'Proses OCR', atau tambah baris kosong secara manual.")



# =====================================================================
# NAVIGASI / MENU UTAMA
# =====================================================================
st.sidebar.title("📚 Menu")
menu = st.sidebar.radio(
    "Pilih fitur:",
    ["📄 Convert CSV ke XLSX", "📸 Ekstrak Tabel dari Foto (OCR Lokal)"],
    key="main_menu"
)

if menu == "📄 Convert CSV ke XLSX":
    render_csv_to_xlsx()
else:
    render_photo_extraction()
